#!/usr/bin/env python3
"""
Majority Vote Analysis Script

For each model type, performs majority vote across 5 dataset ensembles to produce
a final image-level diagnosis. Outputs two XLSX files:
1. confusion_matrices.xlsx - One sheet per model with 4x4 confusion matrix
2. classification_metrics.xlsx - Per-class one-vs-rest metrics + aggregate

Usage:
    python majority_vote_analysis.py
    python majority_vote_analysis.py --models nnunet convnextv2
"""

import os
import sys
import io
import json
import argparse
import logging
from pathlib import Path
from collections import Counter
from typing import Dict, List, Tuple, Optional

import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Fix encoding for Windows console
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('majority_vote_analysis.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Constants
DATASETS = {
    100: 'Liver1',
    101: 'Liver2',
    102: 'Liver3',
    103: 'Liver4',
    104: 'Liver5'
}

DIAGNOSIS_NAMES = ['Healthy', 'Metastatic', 'HCC', 'CHO']
CLASS_LABELS = [0, 1, 2, 3]

# Model definitions with diagnosis level
MODELS_INFO = {
    'nnunet':         'Pixel',
    'nnunet_patch':   'Patch',
    'convnextv2':     'Patch',
    'densenet':       'Patch',
    'efficientnetv2': 'Patch',
    'maxvit':         'Patch',
    'swinv2':         'Patch',
}

# Where to find diagnosis_predictions.json for each model type
def get_diagnosis_path(predictions_dir: Path, model: str, ds_id: int, ds_name: str) -> Path:
    ds_key = f"Dataset{ds_id}_{ds_name}"
    if model == 'nnunet':
        # nnunet stores at dataset level, not in ensemble subfolder
        return predictions_dir / model / ds_key / "diagnosis_predictions.json"
    elif model == 'nnunet_patch':
        return predictions_dir / model / ds_key / "ensemble" / "diagnosis_predictions.json"
    else:
        # Classification models
        return predictions_dir / model / ds_key / "ensemble" / "diagnosis_predictions.json"


def infer_ground_truth(case_id: str) -> Optional[int]:
    """Infer ground truth class from filename prefix."""
    lower = case_id.lower()
    if lower.startswith('cho'):
        return 3
    elif lower.startswith('hcc'):
        return 2
    elif lower.startswith('metastatic'):
        return 1
    return None


def is_intra(case_id: str) -> bool:
    """Check if case is an intra-operative metastatic image."""
    return 'intra' in case_id.lower()


def load_model_predictions(predictions_dir: Path, model: str,
                           dataset_ids: List[int]) -> Dict[str, List[int]]:
    """Load diagnosis predictions from all datasets for a model.

    Returns: {case_id: [diagnosis_from_ds1, diagnosis_from_ds2, ...]}
    """
    case_predictions = {}

    for ds_id in dataset_ids:
        ds_name = DATASETS[ds_id]
        path = get_diagnosis_path(predictions_dir, model, ds_id, ds_name)

        if not path.exists():
            logger.warning(f"  Missing: {path}")
            continue

        with open(path) as f:
            data = json.load(f)

        for case_id, info in data.items():
            diag = info['diagnosis']
            if case_id not in case_predictions:
                case_predictions[case_id] = []
            case_predictions[case_id].append(diag)

    return case_predictions


def majority_vote(predictions: List[int]) -> int:
    """Majority vote across dataset predictions."""
    if not predictions:
        return 0
    counts = Counter(predictions)
    return counts.most_common(1)[0][0]


def compute_ovr_metrics(y_true: np.ndarray, y_pred: np.ndarray, pos_class: int) -> Dict[str, float]:
    """Compute one-vs-rest binary metrics for a given class."""
    tp = np.sum((y_true == pos_class) & (y_pred == pos_class))
    tn = np.sum((y_true != pos_class) & (y_pred != pos_class))
    fp = np.sum((y_true != pos_class) & (y_pred == pos_class))
    fn = np.sum((y_true == pos_class) & (y_pred != pos_class))

    acc = (tp + tn) / (tp + tn + fp + fn) if (tp + tn + fp + fn) > 0 else 0.0
    se = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    sp = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    bacc = (se + sp) / 2.0
    ppv = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    f1 = 2 * ppv * se / (ppv + se) if (ppv + se) > 0 else 0.0

    return {'ACC': acc, 'BACC': bacc, 'SE': se, 'SP': sp, 'F1': f1, 'PPV': ppv}


def write_confusion_matrices(wb: Workbook, results: Dict[str, Tuple],
                             models: List[str]):
    """Write confusion matrix sheets - one per model."""
    from sklearn.metrics import confusion_matrix as sk_cm

    bold = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    sub_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    for model in models:
        if model not in results:
            continue

        y_true, y_pred, case_ids = results[model]
        ws = wb.create_sheet(title=model)

        # Title
        ws.cell(row=1, column=1, value=f"Confusion Matrix - {model}").font = Font(bold=True, size=13)

        # Headers
        ws.cell(row=3, column=1, value="True \\ Predicted").font = bold
        ws['A3'].border = thin_border
        ws['A3'].fill = header_fill

        for j, name in enumerate(DIAGNOSIS_NAMES):
            cell = ws.cell(row=3, column=j + 2, value=name)
            cell.font = bold
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 14

        ws.column_dimensions['A'].width = 22

        # Compute confusion matrix (rows=true, cols=pred)
        cm = sk_cm(y_true, y_pred, labels=CLASS_LABELS)

        for i, name in enumerate(DIAGNOSIS_NAMES):
            cell = ws.cell(row=4 + i, column=1, value=name)
            cell.font = bold
            cell.border = thin_border
            cell.fill = header_fill

            for j in range(4):
                cell = ws.cell(row=4 + i, column=j + 2, value=int(cm[i, j]))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if i == j:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        ws.cell(row=8, column=1, value=f"Total samples: {len(y_true)}").font = Font(italic=True)

        # Metastatic breakdown rows
        ws.cell(row=10, column=1, value="Metastatic Breakdown").font = Font(bold=True, size=11)

        # Headers for breakdown
        for j, name in enumerate(DIAGNOSIS_NAMES):
            cell = ws.cell(row=11, column=j + 2, value=name)
            cell.font = bold
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=11, column=1, value="True \\ Predicted").font = bold
        ws.cell(row=11, column=1).border = thin_border
        ws.cell(row=11, column=1).fill = header_fill

        met_mask = y_true == 1
        intra_mask = np.array([is_intra(c) for c in case_ids])

        subgroups = [
            ("Metastatic", met_mask & ~intra_mask),
            ("Metastatic Intra", met_mask & intra_mask),
            ("Metastatic All", met_mask),
        ]

        for row_offset, (label, mask) in enumerate(subgroups):
            r = 12 + row_offset
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = bold
            cell.border = thin_border
            cell.fill = sub_fill

            sub_pred = y_pred[mask]
            n = int(mask.sum())
            for j, cls in enumerate(CLASS_LABELS):
                count = int(np.sum(sub_pred == cls))
                cell = ws.cell(row=r, column=j + 2, value=count)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if cls == 1:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            # Add N column
            ws.cell(row=11, column=6, value="N").font = bold
            ws.cell(row=11, column=6).border = thin_border
            ws.cell(row=11, column=6).fill = header_fill
            ws.cell(row=11, column=6).alignment = Alignment(horizontal='center')
            cell = ws.cell(row=r, column=6, value=n)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')


def compute_ovr_metrics_filtered(y_true: np.ndarray, y_pred: np.ndarray,
                                  pos_class: int, pos_mask: np.ndarray) -> Dict[str, float]:
    """Compute one-vs-rest metrics where only pos_mask positive cases are counted.

    Sensitivity reflects the subgroup only; specificity uses all non-positive cases.
    """
    tp = np.sum(pos_mask & (y_true == pos_class) & (y_pred == pos_class))
    fn = np.sum(pos_mask & (y_true == pos_class) & (y_pred != pos_class))
    fp = np.sum((y_true != pos_class) & (y_pred == pos_class))
    tn = np.sum((y_true != pos_class) & (y_pred != pos_class))

    acc = (tp + tn) / (tp + tn + fp + fn) if (tp + tn + fp + fn) > 0 else 0.0
    se = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    sp = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    bacc = (se + sp) / 2.0
    ppv = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    f1 = 2 * ppv * se / (ppv + se) if (ppv + se) > 0 else 0.0

    return {'ACC': acc, 'BACC': bacc, 'SE': se, 'SP': sp, 'F1': f1, 'PPV': ppv}


def write_classification_metrics(wb: Workbook, results: Dict[str, Tuple],
                                 models: List[str]):
    """Write per-class and aggregate metric sheets.

    Metastatic class is split into three sheets:
      Metastatic          - non-intra only
      Metastatic Intra    - intra only
      Metastatic All      - combined
    """
    bold = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    metric_names = ['ACC', 'BACC', 'SE', 'SP', 'F1', 'PPV']

    # Sheet definitions: (sheet_name, type)
    # type: 'class' with class_idx, 'met_sub' with subgroup filter, 'aggregate'
    sheet_defs = [
        ('Background',       'class', 0),
        ('Metastatic',       'met_sub', 'non_intra'),
        ('Metastatic Intra', 'met_sub', 'intra'),
        ('Metastatic All',   'met_sub', 'all'),
        ('HCC',              'class', 2),
        ('CHO',              'class', 3),
        ('Aggregate',        'aggregate', None),
    ]

    # Pre-compute standard OVR metrics per class
    all_metrics = {}
    for model in models:
        if model not in results:
            continue
        y_true, y_pred, case_ids = results[model]
        all_metrics[model] = {}
        for cls_idx in [0, 1, 2, 3]:
            all_metrics[model][cls_idx] = compute_ovr_metrics(y_true, y_pred, cls_idx)

        # Metastatic subgroup metrics
        met_mask = y_true == 1
        intra_mask = np.array([is_intra(c) for c in case_ids])
        all_metrics[model]['met_non_intra'] = compute_ovr_metrics_filtered(
            y_true, y_pred, 1, met_mask & ~intra_mask)
        all_metrics[model]['met_intra'] = compute_ovr_metrics_filtered(
            y_true, y_pred, 1, met_mask & intra_mask)
        all_metrics[model]['met_all'] = all_metrics[model][1]  # same as full class 1

    for sheet_name, sheet_type, sheet_key in sheet_defs:
        ws = wb.create_sheet(title=sheet_name)

        ws.cell(row=1, column=1, value=sheet_name).font = Font(bold=True, size=13)

        headers = ['Model', 'Diagnosis Level'] + metric_names
        for j, h in enumerate(headers):
            cell = ws.cell(row=3, column=j + 1, value=h)
            cell.font = bold
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 16
        for col_idx in range(3, 3 + len(metric_names)):
            col_letter = ws.cell(row=3, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 10

        for row_idx, model in enumerate(models):
            if model not in all_metrics:
                continue

            r = 4 + row_idx
            ws.cell(row=r, column=1, value=model).font = bold
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2, value=MODELS_INFO[model]).border = thin_border
            ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')

            if sheet_type == 'class':
                metrics = all_metrics[model][sheet_key]
            elif sheet_type == 'met_sub':
                key_map = {'non_intra': 'met_non_intra', 'intra': 'met_intra', 'all': 'met_all'}
                metrics = all_metrics[model][key_map[sheet_key]]
            else:
                # Aggregate: macro-average across 4 classes
                metrics = {}
                for m in metric_names:
                    vals = [all_metrics[model][c][m] for c in [0, 1, 2, 3]]
                    metrics[m] = np.mean(vals)

            for j, m in enumerate(metric_names):
                cell = ws.cell(row=r, column=j + 3, value=round(metrics[m], 4))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '0.0000'


def main():
    parser = argparse.ArgumentParser(
        description='Cross-dataset majority vote analysis with XLSX output')
    parser.add_argument('--predictions_dir', type=str,
                        default='predictions_external',
                        help='Base directory with model predictions')
    parser.add_argument('--output_dir', type=str,
                        default='predictions_external',
                        help='Output directory for XLSX files')
    parser.add_argument('--models', nargs='+', type=str,
                        default=list(MODELS_INFO.keys()),
                        help='Models to include in analysis')
    parser.add_argument('--datasets', nargs='+', type=int,
                        default=[100, 101, 102, 103, 104],
                        help='Dataset IDs to include')
    parser.add_argument('--skip_substring', nargs='+', type=str, default=[],
                        help='Skip images whose name contains any of these substrings '
                             '(e.g. --skip_substring _400_ _intra_)')
    args = parser.parse_args()

    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is required: pip install openpyxl")
        sys.exit(1)

    predictions_dir = Path(args.predictions_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    skip_subs = [s.lower() for s in args.skip_substring]

    models = [m for m in args.models if m in MODELS_INFO]
    logger.info(f"Majority Vote Analysis")
    logger.info(f"  Models: {models}")
    logger.info(f"  Datasets: {args.datasets}")
    if skip_subs:
        logger.info(f"  Skipping images containing: {skip_subs}")

    # Collect results: model -> (y_true, y_pred, case_ids)
    results = {}

    for model in models:
        logger.info(f"Processing {model}...")
        case_preds = load_model_predictions(predictions_dir, model, args.datasets)

        if not case_preds:
            logger.warning(f"  No predictions found for {model}")
            continue

        y_true_list = []
        y_pred_list = []
        case_id_list = []
        skipped = 0

        for case_id, preds in sorted(case_preds.items()):
            # Skip if case_id contains any of the skip substrings
            if any(s in case_id.lower() for s in skip_subs):
                skipped += 1
                continue

            gt = infer_ground_truth(case_id)
            if gt is None:
                logger.warning(f"  Cannot infer ground truth for: {case_id}")
                continue

            final_diag = majority_vote(preds)
            y_true_list.append(gt)
            y_pred_list.append(final_diag)
            case_id_list.append(case_id)

        if skipped:
            logger.info(f"  Skipped {skipped} images")

        y_true = np.array(y_true_list)
        y_pred = np.array(y_pred_list)
        case_ids = np.array(case_id_list)
        results[model] = (y_true, y_pred, case_ids)

        correct = np.sum(y_true == y_pred)
        logger.info(f"  {model}: {correct}/{len(y_true)} correct "
                     f"({100 * correct / len(y_true):.1f}%)")

    # Write XLSX 1: Confusion Matrices
    wb_cm = Workbook()
    wb_cm.remove(wb_cm.active)  # Remove default sheet
    write_confusion_matrices(wb_cm, results, models)

    cm_path = output_dir / "confusion_matrices.xlsx"
    wb_cm.save(cm_path)
    logger.info(f"Saved confusion matrices: {cm_path}")

    # Write XLSX 2: Classification Metrics
    wb_metrics = Workbook()
    wb_metrics.remove(wb_metrics.active)  # Remove default sheet
    write_classification_metrics(wb_metrics, results, models)

    metrics_path = output_dir / "classification_metrics.xlsx"
    wb_metrics.save(metrics_path)
    logger.info(f"Saved classification metrics: {metrics_path}")

    logger.info("Done.")


if __name__ == '__main__':
    main()
